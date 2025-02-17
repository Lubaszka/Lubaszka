import re
import PyPDF2
import pandas as pd
from pathlib import Path
import os
from datetime import datetime
import time
from tkinter import messagebox
from ttkbootstrap import Style
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk


import re
import PyPDF2
import re
import PyPDF2
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import os


def extract_data_from_pdf(file):
    """Extract relevant data from a PDF file."""
    try:
        pdf_reader = PyPDF2.PdfReader(file)
        text = "\n".join(page.extract_text() or "" for page in pdf_reader.pages)
    except Exception as e:
        return {'error': str(e)}

    data = {
        'ZAHLUNGSDATUM': None,
        'Rechnungsnummer': None,
        'Rechnungsbetrag': None,
        'Nettobetrag': None,
        'MwSt.': None,
        'WV': None,
        'Kommission': None,
        'Rechnungdatum': None,
        'SERVICEDATUM': None,
        'AllText': text,
    }
    
    patterns = {
        'ZAHLUNGSDATUM': r'Datum\s*:\s*(\d{2}\.\d{2}\.\d{4})',
        'Rechnungsnummer': r'R E C H N U N G - Nr\.\:\s*(\d+)',
        'Rechnungsbetrag': r'Gesamt-Betrag €\s*([\d,.]+)',
        'Nettobetrag': r'Netto-Summe €\s*([\d,.]+)',
        'MwSt.': r'19,00 % USt\.\s*€\s*([\d,.]+)',
        'WV': r"(WV\s*[\d,.]+)",
        'Kommission': r"\b(120\d{4})\b",
        'Rechnungdatum': r'Datum\s*:\s*(\d{2}\.\d{2}\.\d{4})',
        'SERVICEDATUM': r'Lieferdatum\s*:\s*(\d{2}\.\d{2}\.\d{4})'
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            data[key] = match.group(1)

    return data
    # Procesujemy inne wzorce
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            data[key] = match.group(1).strip() if key != 'SERVICEDATUM' else f"{match.group(1)} - {match.group(2)}"

    return data

def convert_dates(df):
    """Convert date columns to the format ROKMIESIACDZIEN."""
    for date_column in ['Rechnungdatum', 'ZAHLUNGSDATUM']:
        if date_column in df:
            if date_column == 'ZAHLUNGSDATUM':
                df[date_column] = pd.to_datetime(df[date_column], format='%d.%m.%Y') + timedelta(days=30)
            else:
                df[date_column] = pd.to_datetime(df[date_column], format='%d.%m.%Y')
            df[date_column] = df[date_column].dt.strftime('%Y%m%d')
    return df

def extract_data_with_full_text_from_pdfs(folder_path):
    """Extract data from all PDFs in the given folder."""
    folder = Path(folder_path)
    if not folder.exists():
        print("Podana ścieżka nie istnieje.")
        return None
    
    pdf_data = []

    for pdf_file in folder.glob('*.pdf'):
        try:
            with pdf_file.open('rb') as file:
                data = extract_data_from_pdf(file)
                data['FileName'] = pdf_file.name
                pdf_data.append(data)
        except Exception as e:
            print(f"Błąd podczas odczytu pliku {pdf_file.name}: {e}")

    if not pdf_data:
        print("Nie znaleziono żadnych danych.")
        return None
    
    df = pd.DataFrame(pdf_data)

    df['Rechnungsbetrag'] = df['Rechnungsbetrag'].fillna('0').str.replace('.', '', regex=False).str.replace(',', '.', regex=False).astype(float)
    df['Nettobetrag'] = df['Nettobetrag'].fillna('0').str.replace('.', '', regex=False).str.replace(',', '.', regex=False).astype(float)
    df['MwSt.'] = df['MwSt.'].fillna('0').str.replace('%', '', regex=False).str.replace(',', '.', regex=False).astype(float)
    df['Kommission'] = df['Kommission'].fillna('0').str.replace(',', '.').astype(float)
    
    df = convert_dates(df)
    
    final_df = df[['Rechnungsnummer', 'Rechnungsbetrag', 'Nettobetrag', 'MwSt.', 'WV', 'Kommission', 'Rechnungdatum', 'ZAHLUNGSDATUM', 'SERVICEDATUM', 'AllText']]
    
    output_file = folder / 'W_Invoice.xlsx'
    final_df.to_excel(output_file, index=False)
    print(f"Plik Excel został zapisany jako {output_file}")

    filenames_df = df[['FileName']]
    filename_output_file = folder / 'FileName.xlsx'
    filenames_df.to_excel(filename_output_file, index=True)
    print(f"Plik Excel z nazwami plików został zapisany jako {filename_output_file}")
    
    return final_df

# Example usage:
folder_path = r'C:\Python\Birgels'
w_invoice = extract_data_with_full_text_from_pdfs(folder_path)



# Part 3 - VLookup
if w_invoice is not None:
    AmRest = pd.read_excel(r'C:\Users\lukasz.lubaszka\OneDrive - AmRest\# Finance DE team\AmLocation\AmLocation_SBX_KFC - Active.xlsx')

    # Perform the merge operation
    merged = pd.merge(
        w_invoice,
        AmRest,
        left_on='Kommission',
        right_on='AmRest Number',
        how='left'
    )

    result = merged[['Rechnungsnummer', 'AmRest Number', 'Nettobetrag', 'Rechnungsbetrag', 'MwSt.', 'WV', 'Rechnungdatum', 'ZAHLUNGSDATUM', 'SERVICEDATUM']].copy()
    result['Item name'] = merged.apply(lambda row: ' '.join(row.values.astype(str)), axis=1)

    result.columns = ['Invoice code', 'Orga node', 'Net amount Products', 'Total Gross Amount Invoice', 'Tax amount', 'WV', 'Invoice date', 'Due Date', 'SERVICEDATUM', 'Application Data']

    # Save the result to a new Excel file
    result.to_excel('C:\\Python\\Birgels\\W_Invoice_Import.xlsx', index=False)
    print("Wynik zapisano.")

    # Remove the W_Invoice file
    if os.path.exists('C:\\Python\\Birgels\\W_Invoice.xlsx'):
        os.remove('C:\\Python\\Birgels\\W_Invoice.xlsx')
        print(f'Plik W_Invoice.xlsx został usunięty.')
    else:
        print(f'Plik W_Invoice.xlsx nie istnieje.')


class InvoiceMerger:
    def __init__(self):
        self.invoice_file = 'C:\\Python\\Birgels\\W_Invoice_Import.xlsx'
        self.template_file = r"C:\Python\01.Data\Import_Temple\template.xlsx"

    def load_files(self):
        try:
            self.df_invoice = pd.read_excel(self.invoice_file)
            self.df_template = pd.read_excel(self.template_file)
            print("Pliki załadowane pomyślnie.")
        except FileNotFoundError as e:
            print(f"BŁĄD: Nie znaleziono pliku - {e}")
            raise

    def gather_user_inputs(self, account_date, creation_date):
        self.account_date = account_date
        self.creation_date = creation_date

    def merge_data(self):
        common_columns = ['Invoice code', 'Orga node', 'Invoice date', 'Due Date', 'Total Gross Amount Invoice', 'Tax amount', 'Net amount Products']

              # Check for missing columns
        missing_columns_invoice = [col for col in common_columns if col not in self.df_invoice.columns]
        missing_columns_template = [col for col in common_columns if col not in self.df_template.columns]

        if missing_columns_invoice or missing_columns_template:
            if missing_columns_invoice:
                print(f"BŁĄD: Nie znaleziono kolumn w pliku W_Invoice_Import: {missing_columns_invoice}")
            if missing_columns_template:
                print(f"BŁĄD: Nie znaleziono kolumn w pliku szablonu: {missing_columns_template}")
            return

        # Merge DataFrames
        merged_df = pd.merge(self.df_template, self.df_invoice[common_columns], on=common_columns, how='inner')

        # Add invoice rows
        self.add_invoice_rows(merged_df)

    def add_invoice_rows(self, merged_df):
        invoice_rows = []

        for index, row in self.df_invoice.iterrows():
            invoice_rows.append({
                'Invoice type code': 'INV',
                'Suppliers': 28028,
                'Account date': self.account_date,
                'Units ###': 'EUR',
                'Status': 'ini',
                'Orga level': 'Site',
                'Tax code': 'V1',
                'Ledger Account': 40304000,
                'Item name': 'Birgels',
                'Quantity': '1',
                'Fiscal Year': '2025',
                'Created by': 'patryk.noga@amrest.eu',
                'Legal Company': 'AMRDE',
                'Creation date': self.creation_date,
                'Invoice date': row['Invoice date'],  # Get the correct Invoice date
                'Due Date': row['Due Date'],          # Get the correct Due Date
                '_inv_inv_no_po': 1,
                "Don't send to SAP": 0,
                '_inv_in_noapp': 0,
            })

        df_new_invoices = pd.DataFrame(invoice_rows)
        merged_df = pd.concat([merged_df, df_new_invoices], ignore_index=True)
        self.df_template = merged_df

        print("Nowe wiersze dodane pomyślnie.")
        print(self.df_template.head())

    def save_merged_file(self):
        output_file = r"C:\Python\Birgels\Merged.xlsx"
        self.df_template.to_excel(output_file, index=False)
        return f"Plik został zapisany jako: {output_file}"

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.style = Style(theme='darkly')
        self.root.title("Invoice Merger")
        self.root.geometry("800x600")
        self.root.configure(bg="#ffffff")

        # Load a subtle background image
        self.bg_image = Image.open('C:\\Python\\01.Data\\Import_Temple\\amrest_bg.jpg')
        self.bg_image = self.bg_image.resize((800, 600), Image.LANCZOS)
        self.bg_photo = ImageTk.PhotoImage(self.bg_image)

        self.canvas = tk.Canvas(root, width=800, height=600, bg="#ffffff", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        self.canvas.create_image(0, 0, image=self.bg_photo, anchor="nw")

        self.show_input_form()

    def show_input_form(self):
        for widget in self.canvas.winfo_children():
            widget.destroy()

        input_frame = ttk.Frame(self.canvas, padding=(20, 20))
        input_frame.place(relx=0.5, rely=0.5, anchor="center")

        # Get today's date in YYYYMMDD format
        today = datetime.now().strftime("%Y%m%d")

        ttk.Label(input_frame, text="Account Date (YYYYMMDD):", font=("Arial", 12)).grid(row=0, column=0, padx=10, pady=10)
        self.account_date_entry = ttk.Entry(input_frame)
        self.account_date_entry.insert(0, today)  # Suggest today's date
        self.account_date_entry.grid(row=0, column=1, padx=10, pady=10)

        ttk.Label(input_frame, text="Creation Date (YYYYMMDD):", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=10)
        self.creation_date_entry = ttk.Entry(input_frame)
        self.creation_date_entry.insert(0, today)  # Suggest today's date
        self.creation_date_entry.grid(row=1, column=1, padx=10, pady=10)

        merge_button = ttk.Button(input_frame, text="Merge Invoices", command=self.merge_invoices)
        merge_button.grid(row=2, column=1, padx=10, pady=20)
        merge_button.configure(style='Large.TButton')

    def merge_invoices(self):
        account_date = self.account_date_entry.get()
        creation_date = self.creation_date_entry.get()

        if not account_date or not creation_date:
            messagebox.showerror("Error", "Please fill in all fields with valid data.")
            return

        try:
            merger = InvoiceMerger()
            merger.load_files()
            merger.gather_user_inputs(account_date, creation_date)
            merger.merge_data()
            message = merger.save_merged_file()
            messagebox.showinfo("Success", message)
            self.root.destroy()  # Close the window after successful operation
        except Exception as e:
            messagebox.showerror("Error", f"ERROR: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()


# Additional code for file merging and deletions can follow here...
# Ensure it runs after the GUI app is done.

# Define the file paths to delete
files_to_delete = [
    "C:\\Python\\Birgels\\W_Invoice_Import.xlsx",
    "C:\\Python\\Birgels\\Merged.xlsx"
]

# Load the first Excel file
plik_excel_1 = 'C:\\Python\\Birgels\\W_Invoice_Import.xlsx'
df1 = pd.read_excel(plik_excel_1)

# Add a column with a custom index (starting from 1)
df1['Index'] = range(1, len(df1) + 1)

# Load the second Excel file
plik_excel_2 = 'C:\\Python\\Birgels\\Merged.xlsx'
df2 = pd.read_excel(plik_excel_2)

# Add a column with a custom index (starting from 1)
df2['Index'] = range(1, len(df2) + 1)

# Merge tables based on the 'Index' column
df_polaczone = pd.merge(df1, df2, on='Index', how='left')

# Save the merged table to a new Excel file
df_polaczone.to_excel('C:\\Python\\Birgels\\Import_Birgels.xlsx', index=False)

# Wait for 3 seconds
time.sleep(3)


import re
import PyPDF2
import pandas as pd
from pathlib import Path
import os
import re
import PyPDF2
import pandas as pd
from pathlib import Path
import os
import time
from openpyxl import load_workbook
import os
import zipfile
import numpy as np
import sqlite3
from datetime import datetime



# Usunięcie plików
for file_path in files_to_delete:
    try:
        os.remove(file_path)
        print(f"Usunięto plik: {file_path}")
    except FileNotFoundError:
        print(f"Plik nie znaleziony: {file_path}")
    except Exception as e:
        print(f"Wystąpił błąd podczas usuwania pliku {file_path}: {e}")
import pandas as pd
import tkinter as tk
from tkinter import messagebox


def process_data():
    try:
        # Load Excel file
        df = pd.read_excel('C:\\Python\\Birgels\\Import_Birgels.xlsx', engine='openpyxl')

        # Dodaj nowe kolumny
        df['Description1'] = np.where(df['WV'].notna(), 'KE MAINT contr.repair - TESTTEST', 'KE MAINT repair - TESTTEST')
        df['Description2'] = 'Birgels'
        
        # Konwersja do formatu daty
        df['SERVICEDATUM'] = pd.to_datetime(df['SERVICEDATUM'], format='%d.%m.%Y')
        
        # Formatowanie do 'MM.YY'
        df['SERVICEDATUM'] = df['SERVICEDATUM'].dt.strftime('%m.%y')
        
        df['Description3'] = df['SERVICEDATUM']

        # Combine Description1, Description2, and Description3
        df['Combined_Description'] = df['Description1'] + ',' + df['SERVICEDATUM'].astype(str) + ',' + df['Description2']

        # Remove Description1, Description2, and Description3
        df.drop(columns=['Description1', 'Description2', 'Description3'], inplace=True)

        # Save the result to a new Excel file
        df.to_excel('C:\\Python\\Birgels\\Import_Birgels.xlsx', index=False)

        # Notify the user and close the window
        messagebox.showinfo("Success", "Data processed successfully!")
        root.destroy()  # Close the window after clicking OK

    except KeyError as e:
        messagebox.showerror("Error", f"KeyError: {str(e)}. Ensure all expected columns are present.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Create the main window
root = tk.Tk()
root.title("Birgels")
root.geometry("500x200")  # Set the window size

tk.Label(root, text="Processing data...").pack(pady=20)

tk.Button(root, text="Process Data", command=process_data).pack(pady=20)

# Run the application
root.mainloop()

import pandas as pd
from openpyxl import load_workbook

# Load the DataFrame from Excel
df = pd.read_excel('C:\\Python\\Birgels\\Import_Birgels.xlsx', engine='openpyxl')

column_order = [
    'Invoice type code',
    'Invoice code_x',
    'Suppliers',
    'Invoice date_x',
    'Account date',
    'Units ###',
    'Status',
    'Orga node_x',
    'Scanned image of the invoice',
    'Created by',
    'Legal Company',
    'Combined_Description',
    'Creation date',
    '_inv_inv_no_po',
    "Don't send to SAP",
    'Due Date_x',
    'Delivery date',
    '_inv_in_noapp',
    'Tax code',
    'Tax amount_x',
    'Total Gross Amount Invoice_x',
    'Item name',
    'Quantity',
    'Unit price',
    'Net amount Products_x',
    'Amount (Incl. Tax) Products',
    'Ledger Account',
    'Cost Center',
    'Fiscal Year'
]

# Reorder the DataFrame
df = df[column_order]

# Set the values
df['Cost Center'] = df['Orga node_x']
df['Amount (Incl. Tax) Products'] = df['Total Gross Amount Invoice_x']  # Ensure this is correct
df['Unit price'] = df['Net amount Products_x']
df['Delivery date'] = df['Invoice date_x']

# Uncomment if needed
df['Total Gross Amount Invoice_x'] = df['Total Gross Amount Invoice_x'].where(df['Total Gross Amount Invoice_x'] >= 0, 'CN')



# Define the mapping of old column names to new column names
rename_dict = {
    'Invoice code_x': 'Invoice code',
    'Invoice date_x': 'Invoice date',
    'Orga node_x': 'Orga node',
    'Due Date_x': 'Due date',
    'Tax amount_x': 'Tax amount',
    'Total Gross Amount Invoice_x': 'Total Gross Amount Invoice',
    'Net amount Products_x': 'Net amount Products'
}

# Rename the columns
df.rename(columns=rename_dict, inplace=True)

# Upewnij się, że kolumna jest typu string
df['Orga node'] = df['Orga node'].astype(str)

# Sprawdź i ustaw wartości na podstawie warunków
df.loc[df['Orga node'].str.startswith('1208'), 'Orga node'] = '1202058'
df.loc[df['Orga node'].str.startswith('1203'), 'Orga node'] = '1202088'



import os
import time
import zipfile
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk

# Save the result to a new Excel file
output_path = 'C:\\Python\\Birgels\\Organized_Import_Birgels.xlsx'
df.to_excel(output_path, index=False)

# Adjust column widths
wb = load_workbook(output_path)
ws = wb.active

for column in ws.columns:
    max_length = 0
    for cell in column:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except Exception as e:
            print(f"Error processing cell {cell}: {e}")
    adjusted_width = max_length + 2
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Save the workbook again
wb.save(output_path)

# Load tables from Excel files
table1 = pd.read_excel('C:\\Python\\Birgels\\FileName.xlsx')
table2 = pd.read_excel(output_path)

# Add new column to table2
table2['Scanned image of the invoice'] = table1['FileName']

# Wait for a short duration
time.sleep(1)

# Paths to delete
files_to_delete = [
    "C:\\Python\\Birgels\\Import_Birgels.xlsx",
    "C:\\Python\\Birgels\\FileName.xlsx",
]

# Save modified table to a new Excel file
table2.to_excel(output_path, index=False)

# Delete unnecessary files
for file_path in files_to_delete:
    try:
        os.remove(file_path)
        print(f"Deleted file: {file_path}")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error deleting file {file_path}: {e}")

# Wait for a short duration
time.sleep(1)

# Open the Excel file
os.startfile(output_path)

def zip_files(source_folder, output_zip):
    # Create a ZIP file
    with zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(source_folder):
            file_path = os.path.join(source_folder, file)
            if os.path.isfile(file_path) and file.endswith(('.pdf', '.xls', '.xlsx')):
                zipf.write(file_path, os.path.relpath(file_path, source_folder))

if __name__ == "__main__":
    # Define folder location and ZIP file name
    source_folder = r'C:\Python\Birgels'  # Change to your location
    output_zip = r'C:\Python\Birgels\import.zip'  # ZIP file name

    zip_files(source_folder, output_zip)
    print(f"PDF and Excel files from {source_folder} (excluding subfolders) have been archived to {output_zip}.")

# Data processing function
def process_data():
    total_records = len(df)
    empty_values = df.isnull().sum()
    empty_columns = empty_values[empty_values > 0]

    report_message = "\n--------  Data Validation Report  --------\n"
    report_message += f"Total records: {total_records}\n"

    if not empty_columns.empty:
        report_message += "Count of empty values in columns:\n"
        empty_columns_sorted = empty_columns.sort_values()
        for column, count in empty_columns_sorted.items():
            report_message += f"{column}: {count}\n"
    else:
        report_message += "No empty values in columns.\n"

    # Clear previous report and display the new report
    report_text.delete(1.0, tk.END)  # Clear the text area
    report_text.insert(tk.END, report_message)  # Insert the new report

def close_program():
    root.destroy()  # Close the application

# Create the main window
root = tk.Tk()
root.title("Birgels")
root.geometry("800x500")  # Adjust window size for report display

tk.Label(root, text="Processing data...").pack(pady=20)
tk.Button(root, text="Process Data", command=process_data).pack(pady=10)
tk.Button(root, text="End", command=close_program).pack(pady=10)  # Button to close the program

# Create a Text widget to display the report
report_text = tk.Text(root, wrap=tk.WORD, height=10, font=("Helvetica", 15))  # Set font size
report_text.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

# Run the application
root.mainloop()

